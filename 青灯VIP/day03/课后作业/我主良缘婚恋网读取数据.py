'''
功能：抓取http://www.7799520.com/jiaoyou.html 网站的数据
通过抓包工具，分析得到：
    1、查看网页源代码，里面没有数据，则可能是通过js动态加载的。
    2、通过抓包，得到数据保存在http://www.7799520.com/api/user/pc/list/search?marry=1&page=2  这个js里面
    3、这个请求是一个get请求，构造相应参数，即可抓取前10页数据
    4、将数据保存到excel文件中，利用openpyxl库
'''

import json
import requests
from openpyxl import Workbook  # 导入第三方库，操作Excel
# 插入Excel数据使用，累加行号，定义一个全局变量
global row
row = 2


def get_response(page):
    url = f'http://www.7799520.com/api/user/pc/list/search?marry=1&page={page}'
    headers = {
        'User-Agent': "Mozilla/5.0 (Windows NT 6.1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/72.0.3626.109 Safari/537.36"
    }
    return requests.get(url, headers=headers)


def generate_excel(ws, dic_data):
    '''
    :param ws: 创建的工作簿实例化对象
    :param dic_data: 字典数据
    :return: None
    '''
    global row
    ws['A1']='用户编号'
    ws['B1']='省份'
    ws['C1']='城市'
    ws['D1']='身高'
    ws['E1']='学历'
    ws['F1']='姓名'
    ws['G1']='独白'
    ws['H1']='出生年'
    ws['I1']='性别'
    ws['J1']='薪资'
    ws['K1']='婚否'

    col = 1  # 定义列序号，每次迭代都需要重复，故不需要定义为全局变量
    for item in dic_data['data']['list']:
        ws.cell(row=row, column=col, value=item['userid'])
        ws.cell(row=row, column=col+1, value=item['province'])
        ws.cell(row=row, column=col+2, value=item['city'])
        ws.cell(row=row, column=col+3, value=item['height'])
        ws.cell(row=row, column=col+4, value=item['education'])
        ws.cell(row=row, column=col+5, value=item['username'])
        ws.cell(row=row, column=col+6, value=item['monolog'])
        ws.cell(row=row, column=col+7, value=item['birthdayyear'])
        ws.cell(row=row, column=col+8, value=item['gender'])
        ws.cell(row=row, column=col+9, value=item['salary'])
        ws.cell(row=row, column=col+10, value=item['marry'])
        row += 1


if __name__ == '__main__':
    wb = Workbook()
    ws = wb.active
    # 获取前10页的数据
    for page in range(1, 20):
        res = get_response(page)
        # 返回的数据就是json格式，所以不需要转换
        infos = res.json()  # 是一个列表，列表中又有字典
        generate_excel(ws, infos)
    # 所有的数据插入完毕后，保存并关闭文件
    wb.save('结果.xlsx')
    wb.close()


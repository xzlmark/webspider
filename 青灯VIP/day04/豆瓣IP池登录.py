'''
https://movie.douban.com/j/new_search_subjects?sort=U&range=0,10&tags=&start=9979

'''
import random
import json
import requests
from openpyxl import Workbook  # 导入第三方库，操作Excel
global row
row = 2
def get_ua():
    user_agent_list = [
        "Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.1 (KHTML, like Gecko) Chrome/22.0.1207.1 Safari/537.1",
        "Mozilla/5.0 (X11; CrOS i686 2268.111.0) AppleWebKit/536.11 (KHTML, like Gecko) Chrome/20.0.1132.57 Safari/536.11",
        "Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/536.6 (KHTML, like Gecko) Chrome/20.0.1092.0 Safari/536.6",
        "Mozilla/5.0 (Windows NT 6.2) AppleWebKit/536.6 (KHTML, like Gecko) Chrome/20.0.1090.0 Safari/536.6",
        "Mozilla/5.0 (Windows NT 6.2; WOW64) AppleWebKit/537.1 (KHTML, like Gecko) Chrome/19.77.34.5 Safari/537.1",
        "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/536.5 (KHTML, like Gecko) Chrome/19.0.1084.9 Safari/536.5",
        "Mozilla/5.0 (Windows NT 6.0) AppleWebKit/536.5 (KHTML, like Gecko) Chrome/19.0.1084.36 Safari/536.5",
        "Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/536.3 (KHTML, like Gecko) Chrome/19.0.1063.0 Safari/536.3",
        "Mozilla/5.0 (Windows NT 5.1) AppleWebKit/536.3 (KHTML, like Gecko) Chrome/19.0.1063.0 Safari/536.3",
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_8_0) AppleWebKit/536.3 (KHTML, like Gecko) Chrome/19.0.1063.0 Safari/536.3",
        "Mozilla/5.0 (Windows NT 6.2) AppleWebKit/536.3 (KHTML, like Gecko) Chrome/19.0.1062.0 Safari/536.3",
        "Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/536.3 (KHTML, like Gecko) Chrome/19.0.1062.0 Safari/536.3",
        "Mozilla/5.0 (Windows NT 6.2) AppleWebKit/536.3 (KHTML, like Gecko) Chrome/19.0.1061.1 Safari/536.3",
        "Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/536.3 (KHTML, like Gecko) Chrome/19.0.1061.1 Safari/536.3",
        "Mozilla/5.0 (Windows NT 6.1) AppleWebKit/536.3 (KHTML, like Gecko) Chrome/19.0.1061.1 Safari/536.3",
        "Mozilla/5.0 (Windows NT 6.2) AppleWebKit/536.3 (KHTML, like Gecko) Chrome/19.0.1061.0 Safari/536.3",
        "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/535.24 (KHTML, like Gecko) Chrome/19.0.1055.1 Safari/535.24",
        "Mozilla/5.0 (Windows NT 6.2; WOW64) AppleWebKit/535.24 (KHTML, like Gecko) Chrome/19.0.1055.1 Safari/535.24"
    ]
    return {
        'User-Agent': random.choice(user_agent_list)
    }

def get_proxy():
    ip_port = ''
    url = 'http://127.0.0.1:5010/get/'
    try:
        res = requests.get(url, timeout = 10)
        ip_port = res.text
    except:
        ip_port = '127.0.0.1:8080'
    print(ip_port)
    return {
        'http': ip_port,
        'https': ip_port
    }


def generate_excel(ws, dic_data):
    '''
    :param ws: 创建的工作簿实例化对象
    :param dic_data: 字典数据
    :return: None
    '''
    global row
    ws['A1']='导演'
    ws['B1']='评分'
    ws['C1']='格式'
    ws['D1']='星级'
    ws['E1']='电影名'
    ws['F1']='详情网址'
    ws['G1']='主演'
    ws['H1']='封面'
    ws['I1']='ID'
    ws['J1']='格式'

    col = 1  # 定义列序号，每次迭代都需要重复，故不需要定义为全局变量
    print(dic_data)
    for item in dic_data['data']:
        ws.cell(row=row, column=col, value=str(item['directors']))
        ws.cell(row=row, column=col+1, value=item['rate'])
        ws.cell(row=row, column=col+2, value=item['cover_x'])
        ws.cell(row=row, column=col+3, value=item['star'])
        ws.cell(row=row, column=col+4, value=item['title'])
        ws.cell(row=row, column=col+5, value=item['url'])
        ws.cell(row=row, column=col+6, value=str(item['casts']))
        ws.cell(row=row, column=col+7, value=item['cover'])
        ws.cell(row=row, column=col+8, value=item['id'])
        ws.cell(row=row, column=col+9, value=item['cover_y'])
        row += 1


if __name__ == '__main__':
    wb = Workbook()
    ws = wb.active
    url = 'https://movie.douban.com/j/new_search_subjects?sort=U&range=0,10&tags=&start={0}'
    for i in range(0, 100,20):
        proxy = get_proxy()
        res = requests.get(url.format(i), headers=proxy, proxies=proxy,verify=False)
        infos = res.json()  # res.json()返回的数据就是dict格式
        generate_excel(ws, infos)
    # 所有的数据插入完毕后，保存并关闭文件
    wb.save('结果.xlsx')
    wb.close()
